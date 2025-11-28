# udise_data_generator_complete_final_fixed.py
import streamlit as st
import pandas as pd
import numpy as np
import os
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="UDISE Data Generator", layout="wide")

# -------------------------
# Helpers
# -------------------------
def safe_numeric_sum(df: pd.DataFrame, cols: List[str]) -> pd.Series:
    """Sum columns coercing missing / non-numeric to 0."""
    series_list = []
    for c in cols:
        if c in df.columns:
            series_list.append(pd.to_numeric(df[c], errors="coerce").fillna(0))
        else:
            series_list.append(pd.Series([0] * len(df), index=df.index))
    if not series_list:
        return pd.Series([0] * len(df), index=df.index)
    return sum(series_list)

def to_excel_bytes_styled(df: pd.DataFrame, header_fill_color="0070C0") -> bytes:
    """Write df to an excel file in-memory with header styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "udise_extract"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # style header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # apply border to data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # adjust widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
            except:
                val = ""
            if val:
                max_length = max(max_length, len(val))
        ws.column_dimensions[column].width = min(50, (max_length + 2))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# -------------------------
# Translations (basic)
# -------------------------
TRANSLATIONS = {
    "en": {
        "title": "UDISE Data Generator",
        "upload": "Upload School Master List (Excel or CSV)",
        "preview": "Preview Uploaded Data",
        "udise_col": "Select UDISE Column",
        "udise_input": "Enter one or more UDISE Codes (comma or newline separated)",
        "select_columns": "Select Columns to Include in Output",
        "generate": "Generate Output",
        "filters": "Filters",
        "create_calc": "Create Custom Fields (Optional)",
        "calc_type": "Choose Calculation Type",
        "sum": "Sum",
        "diff": "Difference (A - B)",
        "avg": "Average",
        "custom": "Custom Formula (simple)",
        "new_field": "Enter new field name (no spaces recommended)",
        "add_field": "Add Calculated Field",
        "preset_formulas": "Preset Formulas",
        "download": "⬇ Download Excel Output",
        "no_file": "Please upload a master list file to proceed.",
        "no_udise": "Please enter at least one UDISE Code.",
        "no_matches": "⚠ No matching UDISE codes found.",
        "found_matches": "✔ Found {n} matching records",
        "apply_filters": "Apply Filters",
        "save_preset": "Save current computed fields as preset",
        "preset_saved": "Preset saved in session.",
        "preset_applied": "Applied preset {k}"
    },
    "ta": {
        "title": "UDISE தரவு உருவாக்கி",
        "upload": "பள்ளி மாஸ்டர் பட்டியலை (Excel அல்லது CSV) பதிவேற்றுங்கள்",
        "preview": "பதிவேற்றப்பட்ட தரவின் முன்னோட்டம்",
        "udise_col": "UDISE பத்து தேர்ந்தெடுக்கவும்",
        "udise_input": "ஒரு அல்லது அதற்கு மேற்பட்ட UDISE குறியீடுகளை உள்ளிடுங்கள் (கமா அல்லது புதிய வரியில்)",
        "select_columns": "ஏற்றுமதிக்கக் காணிக்கைகள் தேர்வு செய்க",
        "generate": "வெளியீட்டை உருவாக்கு",
        "filters": "வடிகட்டல்கள்",
        "create_calc": "செயல்படுத்தப்பட்ட களங்களை உருவாக்கு (இருப்பினால்)",
        "calc_type": "கணக்கீட்டு வகையை தேர்ந்தெடுக்கவும்",
        "sum": "கூட்டல்",
        "diff": "வித்தியாசம் (A - B)",
        "avg": "சராசா",
        "custom": "தனிப்பயன் சூத்திரம் (எளிது)",
        "new_field": "புதிய களப் பெயரை உள்ளிடவும் (வெற்று இடங்கள் தவிர்க்கவும்)",
        "add_field": "கணக்கீட்டுப் புலம் சேர்க்கவும்",
        "preset_formulas": "முன்னிருப்பு சூத்திரங்கள்",
        "download": "⬇ Excel பதிவிறக்கம்",
        "no_file": "தொடர உங்கள் மாஸ்டர் பட்டியலை பதிவேற்றவும்.",
        "no_udise": "அனைத்து குறைந்தது ஒரு UDISE குறியீட்டை உள்ளிடவும்.",
        "no_matches": "⚠ பொருந்தக்கூடிய UDISE குறியீடுகள் இல்லை.",
        "found_matches": "✔ {n} பொருந்தக்கூடிய பதிவுகள் காணப்பட்டன",
        "apply_filters": "வடிகட்டல்கள் சமர்ப்பி",
        "save_preset": "தற்போதைய கணக்கீட்டுச் புலங்களை மூலமாக சேமிக்கவும்",
        "preset_saved": "முன்னிருப்பு செஷனில் சேமிக்கப்பட்டது.",
        "preset_applied": "முன்னிருப்பு {k} பொருந்தியது"
    }
}

# -------------------------
# Start UI
# -------------------------
st.title("🏫 UDISE Data Generator")

lang_choice = st.radio("Language / மொழி", ("English", "தமிழ்"), horizontal=True)
lang = "en" if lang_choice == "English" else "ta"
tr = TRANSLATIONS[lang]

st.header(tr["title"])

# Sidebar filters header
st.sidebar.header(tr["filters"])

# Session inits
if "formula_presets" not in st.session_state:
    st.session_state["formula_presets"] = {}  # name -> list of fields
if "extra_fields" not in st.session_state:
    st.session_state["extra_fields"] = []  # fields that have been created (preset or user-created) -> shown in dropdown
if "created_fields" not in st.session_state:
    st.session_state["created_fields"] = {}  # name -> metadata for user-created calcs
if "selected_columns" not in st.session_state:
    st.session_state["selected_columns"] = []  # persistent selection state

# -------------------------
# Load Master File (URL → Local → Upload)
# -------------------------

import requests
from io import BytesIO

st.subheader("Master Data Source")

MASTER_URL = "https://d3ijhv7dn0xr3b.cloudfront.net/10684.csv"

df_master = None
source_used = None

# -------------------------------------------
# 1️⃣ Try loading from online master URL first
# -------------------------------------------
try:
    st.write("Fetching default master file from online source...")
    response = requests.get(MASTER_URL, timeout=10)

    if response.status_code == 200:
        data = BytesIO(response.content)

        if MASTER_URL.lower().endswith(".csv"):
            df_master = pd.read_csv(data, dtype=str)
        elif MASTER_URL.lower().endswith(".xls"):
            df_master = pd.read_excel(data, engine="xlrd", dtype=str)
        else:
            df_master = pd.read_excel(data, engine="openpyxl", dtype=str)

        source_used = f"Online URL: {MASTER_URL}"
        st.success(f"✔ Loaded master file from URL")

    else:
        st.warning(f"⚠ URL returned status code: {response.status_code}")

except Exception as e:
    st.warning(f"⚠ Could not load from online URL: {e}")


# -------------------------------------------
# 2️⃣ Try local default master files (fallback)
# -------------------------------------------
if df_master is None:
    default_files = ["master.xlsx", "master.xls", "master.csv"]

    for f in default_files:
        if os.path.exists(f):
            try:
                if f.endswith(".csv"):
                    df_master = pd.read_csv(f, dtype=str)
                elif f.endswith(".xls"):
                    df_master = pd.read_excel(f, engine="xlrd", dtype=str)
                else:
                    df_master = pd.read_excel(f, engine="openpyxl", dtype=str)

                source_used = f"Local file: {f}"
                st.success(f"✔ Loaded default master file: {f}")
                break

            except Exception as e:
                st.warning(f"⚠ Found {f} but could not load it: {e}")


# -------------------------------------------
# 3️⃣ Upload option always overrides previous
# -------------------------------------------
st.subheader("Optional: Upload master file to override default")

uploaded_file = st.file_uploader(
    "Upload Excel/CSV",
    type=["xlsx", "xls", "csv"]
)

if uploaded_file is not None:
    try:
        fname = uploaded_file.name.lower()
        if fname.endswith(".csv"):
            df_master = pd.read_csv(uploaded_file, dtype=str)
        elif fname.endswith(".xls"):
            df_master = pd.read_excel(uploaded_file, engine="xlrd", dtype=str)
        else:
            df_master = pd.read_excel(uploaded_file, engine="openpyxl", dtype=str)

        source_used = f"Uploaded file: {uploaded_file.name}"
        st.success(f"✔ Using uploaded master file: {uploaded_file.name}")

    except Exception as e:
        st.error(f"❌ Error reading uploaded file: {e}")
        st.stop()


# -------------------------------------------
# Final fail-safe
# -------------------------------------------
if df_master is None:
    st.error("❌ No master data available. Please upload a file.")
    st.stop()

# Normalize columns
df_master.columns = df_master.columns.str.strip()

# st.info(f"📌 Using master data from: **{source_used}**")


# --- COERCE class gender columns to numeric early ---
# This is the fix: convert ClassN_Boys/Girls/Transgen -> numeric with fillna(0)
import re
for col in df_master.columns:
    if re.match(r"(?i)^Class\d+_(Boys|Girls|Transgen)$", col):  # case-insensitive match
        df_master[col] = pd.to_numeric(df_master[col], errors="coerce").fillna(0)

# Working copy
df = df_master.copy()

# Sidebar filters - detect available columns for each filter key
filter_cols_candidates = {
    "District": ["District", "district", "DISTRICT", "DISTRICT_NAME"],
    "Block": ["Block", "block", "BLOCK", "BlockName"],
    "Education District": ["Education District", "EducationDistrict", "EDU_DIST", "EDUCATION_DISTRICT"],
    "School Type": ["School Type", "SchoolType", "Type", "SCHOOL_TYPE","School_Type"],
    "Management": ["Management", "management", "MANAGEMENT"],
    "Management Type": ["Management Type", "ManagementType",""],
    "Category": ["Category", "category", "CATEGORY"],
    "Category Type": ["Category Type", "CategoryType","Category_Type"]
}

def find_col(candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

selected_filters = {}
with st.sidebar.form("filters_form"):
    st.write("Filter by (optional):")
    for key, candidates in filter_cols_candidates.items():
        col = find_col(candidates)
        if col:
            options = sorted(df[col].dropna().astype(str).unique().tolist())
            chosen = st.multiselect(f"{key}", options=options, key=f"filter_{key}")
            if chosen:
                selected_filters[col] = chosen
    apply_filters = st.form_submit_button(tr["apply_filters"])

# Apply filters if any
if selected_filters:
    mask = pd.Series([True] * len(df))
    for col, vals in selected_filters.items():
        mask = mask & df[col].astype(str).isin(vals)
    df = df[mask]

# UDISE column auto-detect
udise_candidates = ["UDISE", "UDISE Code", "UDISE_Code", "udise", "udise_code", "UDISECODE"]
udise_col = None
for c in udise_candidates:
    if c in df.columns:
        udise_col = c
        break
if not udise_col:
    udise_col = st.selectbox(tr["udise_col"], options=list(df.columns))

# UDISE input
udise_input = st.text_area(tr["udise_input"], height=80)
udise_list = []
if udise_input:
    udise_list = [u.strip() for u in udise_input.replace("\r", "\n").replace(",", "\n").split("\n") if u.strip()]

# Apply UDISE filter
if udise_list:
    df = df[df[udise_col].astype(str).isin(udise_list)]
else:
    st.warning(tr["no_udise"])
# --- Maintain user-given UDISE order ---
try:
    df[udise_col] = df[udise_col].astype(str)
    df = df.set_index(udise_col).loc[udise_list].reset_index()
except Exception as e:
    st.warning(f"Some UDISE codes not found or ordering issue: {e}")

# Create helper to actually build preset fields on demand
def build_class_totals(target_df):
    """Create Class1_Total ... Class12_Total in the given dataframe object (in place)."""
    # Ensure numeric coercion again in case filtered df has string values
    import re
    for col in target_df.columns:
        if re.match(r"(?i)^Class\d+_(Boys|Girls|Transgen)$", col):
            target_df[col] = pd.to_numeric(target_df[col], errors="coerce").fillna(0)

    for i in range(1, 13):
        members = [f"Class{i}_Boys", f"Class{i}_Girls", f"Class{i}_Transgen"]
        target_df[f"Class{i}_Total"] = safe_numeric_sum(target_df, members)

def build_enrollment_presets(target_df):
    """Create Enrollment_1_5, Enrollment_6_8, Enrollment_9_10, Enrollment_11_12, Total_Enrollment"""
    # Ensure class totals exist (they will be zeros if members missing)
    for i in range(1, 13):
        if f"Class{i}_Total" not in target_df.columns:
            members = [f"Class{i}_Boys", f"Class{i}_Girls", f"Class{i}_Transgen"]
            target_df[f"Class{i}_Total"] = safe_numeric_sum(target_df, members)

    target_df["Enrollment_1_5"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(1,6)])
    target_df["Enrollment_6_8"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(6,9)])
    target_df["Enrollment_9_10"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(9,11)])
    target_df["Enrollment_11_12"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(11,13)])
    target_df["Total_Enrollment"] = safe_numeric_sum(target_df, [f"Class{i}_Total" for i in range(1,13)])

# -------------------------
# Preset / Ensure Buttons - create fields only when user clicks
# -------------------------
st.markdown("---")
st.subheader(tr["preset_formulas"])

col1, col2 = st.columns(2)
with col1:
    if st.button("Ensure: Class1-12 Totals (creates Class1_Total .. Class12_Total)"):
        build_class_totals(df)
        # register these fields as available in dropdown (do NOT auto-select)
        for i in range(1,13):
            cname = f"Class{i}_Total"
            if cname not in st.session_state["extra_fields"]:
                st.session_state["extra_fields"].append(cname)
        st.success("Class totals created and added to dropdown (not auto-selected).")

with col2:
    if st.button("Ensure: Enrollment Presets (1-5,6-8,9-10,11-12,Total)"):
        # ensure class totals exist first
        build_class_totals(df)
        build_enrollment_presets(df)
        preset_names = ["Enrollment_1_5", "Enrollment_6_8", "Enrollment_9_10", "Enrollment_11_12", "Total_Enrollment"]
        for pname in preset_names:
            if pname not in st.session_state["extra_fields"]:
                st.session_state["extra_fields"].append(pname)
        st.success("Enrollment presets created and added to dropdown (not auto-selected).")

# -------------------------
# Calculated fields UI (user created) - add to df and to extra_fields but do NOT auto-select
# -------------------------
st.markdown("---")
st.subheader(tr["create_calc"])

# Determine numeric candidates
numeric_candidates = []
for c in df.columns:
    tmp = pd.to_numeric(df[c], errors="coerce")
    if not tmp.isnull().all():
        numeric_candidates.append(c)

# ensure class totals in numeric candidates if already created
for i in range(1,13):
    cname = f"Class{i}_Total"
    if cname in df.columns and cname not in numeric_candidates:
        numeric_candidates.append(cname)

calc_type = st.selectbox(tr["calc_type"], [tr["sum"], tr["diff"], tr["avg"], tr["custom"]])

if calc_type == tr["diff"]:
    col_a = st.selectbox("Column A", options=numeric_candidates, key="diffA")
    col_b = st.selectbox("Column B", options=numeric_candidates, key="diffB")
    new_field_name = st.text_input(tr["new_field"], key="diff_name")
elif calc_type in (tr["sum"], tr["avg"]):
    cols_to_use = st.multiselect("Select numeric columns", options=numeric_candidates, key="sum_cols")
    new_field_name = st.text_input(tr["new_field"], key="sum_name")
else:
    st.caption("Use column names and operators (+, -, *, /, parentheses). Example: (Class1_Total + Class2_Total) / Total_Enrollment")
    custom_formula = st.text_input("Enter custom formula", key="custom_formula")
    new_field_name = st.text_input(tr["new_field"], key="custom_name")

if st.button(tr["add_field"]):
    if not new_field_name:
        st.error("Enter a valid new field name.")
    else:
        try:
            if calc_type == tr["diff"]:
                a = pd.to_numeric(df[col_a], errors="coerce").fillna(0)
                b = pd.to_numeric(df[col_b], errors="coerce").fillna(0)
                df[new_field_name] = a - b
                meta = ("diff", (col_a, col_b))
            elif calc_type == tr["sum"]:
                if not cols_to_use:
                    st.error("Select at least one column to sum.")
                    raise RuntimeError("no cols")
                df[new_field_name] = safe_numeric_sum(df, cols_to_use)
                meta = ("sum", cols_to_use)
            elif calc_type == tr["avg"]:
                if not cols_to_use:
                    st.error("Select at least one column to average.")
                    raise RuntimeError("no cols")
                df[new_field_name] = safe_numeric_sum(df, cols_to_use) / len(cols_to_use)
                meta = ("avg", cols_to_use)
            else:
                expr = custom_formula.strip()
                env = {c: pd.to_numeric(df[c], errors="coerce").fillna(0) for c in df.columns}
                df[new_field_name] = eval(expr, {"__builtins__": {}}, env)
                meta = ("custom", expr)

            # register as available field (but DO NOT auto-add to selected columns)
            if new_field_name not in st.session_state["extra_fields"]:
                st.session_state["extra_fields"].append(new_field_name)
            # persist metadata so we can rebuild on filtered df before export
            st.session_state["created_fields"][new_field_name] = {"type": meta[0], "definition": meta[1]}

            st.success(f"Field '{new_field_name}' created and added to dropdown (not auto-selected).")
        except Exception as e:
            st.error(f"Error creating field: {e}")

# -------------------------
# Preset saves / apply (session)
# -------------------------
st.markdown("---")
st.subheader("Presets (session)")

preset_name = st.text_input("Preset name (optional for save)")

if st.button(tr["save_preset"]):
    name = preset_name.strip() or f"preset_{len(st.session_state['formula_presets'])+1}"
    # Save the list of extra_fields as this preset (session-only)
    st.session_state["formula_presets"][name] = st.session_state["extra_fields"].copy()
    st.success(tr["preset_saved"])

if st.session_state["formula_presets"]:
    st.write("Saved presets (session):")
    for k, v in st.session_state["formula_presets"].items():
        if st.button(f"Apply preset: {k}"):
            for f in v:
                if f not in st.session_state["extra_fields"]:
                    st.session_state["extra_fields"].append(f)
            st.success(tr["preset_applied"].format(k=k))

# -------------------------
# Column selector (render after presets & calculated fields)
# -------------------------
st.markdown("---")
st.subheader(tr["select_columns"])

# Re-evaluate available columns now that new ones may have been created or registered
all_columns = list(df.columns) + [f for f in st.session_state["extra_fields"] if f not in df.columns]

# Deduplicate preserving order
seen = set()
available_columns = []
for c in all_columns:
    if c not in seen:
        available_columns.append(c)
        seen.add(c)

# Keep previous selections if still available
default_sel = [c for c in st.session_state["selected_columns"] if c in available_columns]
selected_columns = st.multiselect(
    tr["select_columns"],
    options=available_columns,
    default=default_sel,
    key="ui_selected_columns"
)

# Save selection back to session_state
st.session_state["selected_columns"] = selected_columns

# -------------------------
# Generate output & download
# -------------------------
st.markdown("---")
if st.button(tr["generate"]):
    if df.empty:
        st.warning(tr["no_matches"])
    else:
        # Recreate preset/class totals & user-created calculated fields on the current filtered df
        # 1) If class totals were registered (user clicked Ensure), build them
        class_total_names = [f"Class{i}_Total" for i in range(1,13)]
        if any(name in st.session_state["extra_fields"] for name in class_total_names):
            build_class_totals(df)

        # 2) If enrollment presets were registered, (re)build them
        enroll_preset_names = ["Enrollment_1_5", "Enrollment_6_8", "Enrollment_9_10", "Enrollment_11_12", "Total_Enrollment"]
        if any(name in st.session_state["extra_fields"] for name in enroll_preset_names):
            build_enrollment_presets(df)

        # 3) Recreate user-created fields from metadata
        for fname, meta in st.session_state["created_fields"].items():
            if meta["type"] == "diff":
                a, b = meta["definition"]
                df[fname] = pd.to_numeric(df.get(a, pd.Series(0, index=df.index)), errors="coerce").fillna(0) - pd.to_numeric(df.get(b, pd.Series(0, index=df.index)), errors="coerce").fillna(0)
            elif meta["type"] == "sum":
                df[fname] = safe_numeric_sum(df, meta["definition"])
            elif meta["type"] == "avg":
                df[fname] = safe_numeric_sum(df, meta["definition"]) / max(1, len(meta["definition"]))
            elif meta["type"] == "custom":
                env = {c: pd.to_numeric(df.get(c, 0), errors="coerce").fillna(0) for c in df.columns}
                try:
                    df[fname] = eval(meta["definition"], {"__builtins__": {}}, env)
                except Exception:
                    df[fname] = pd.Series(0, index=df.index)

        # Validate selected columns
        valid_selected = [c for c in st.session_state["selected_columns"] if c in df.columns]
        missing = [c for c in st.session_state["selected_columns"] if c not in df.columns]
        if missing:
            st.error(f"The following selected fields are missing from the dataset: {missing}. They may not have been created. Try clicking the Ensure buttons or recreate calculated fields.")
        elif not valid_selected:
            st.error("No valid columns selected for output.")
        else:
            out_df = df[valid_selected].copy()

            st.success(tr["found_matches"].format(n=len(out_df)))
            st.dataframe(out_df.head(50))

            excel_bytes = to_excel_bytes_styled(out_df)
            csv_bytes = out_df.to_csv(index=False).encode("utf-8")

            filename_base = "UDISE_Filtered_Output"
            if lang == "ta":
                filename_base = "UDISE_வெளியீடு"

            st.download_button(tr["download"], data=excel_bytes, file_name=filename_base + ".xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            # -------------------------------------------
            # Provide COPY OUTPUT option
            # -------------------------------------------
            st.markdown("### 📋 Copy Output")

            # Convert output DF to TSV (Excel/Google Sheets friendly)
            copy_text = out_df.to_csv(sep="\t", index=False)

            st.text_area(
            "Copy the entire output (Ctrl + A → Ctrl + C):",
             copy_text,
             height=250
             )
            st.download_button("⬇ Download CSV", data=csv_bytes, file_name=filename_base + ".csv", mime="text/csv")
            st.info("Excel has formatted headers (blue bold) and borders.")

# Footer
st.markdown("---")
st.caption("Built with ❤️ — if some class columns differ from ClassN_Boys/Girls/Transgen, give exact names and I'll adapt.")
